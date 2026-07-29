from __future__ import annotations

from io import BytesIO
import re
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import PoServer, settings
from app.domains.channels.service import ChannelService
from app.integrations.sap_po.client import build_session, endpoint
from app.integrations.sap_po.errors import SapPoConfigurationError, SapPoConnectionError


FROM_PATTERN = re.compile(r"\bFROM\s+([A-Za-z0-9_.$]+)", re.IGNORECASE)
COMPANY_PATTERN = re.compile(
    r"\bCOMPANY_(?:CD|CO)\s+IN\s*\(([^)]+)\)",
    re.IGNORECASE,
)
QUOTED_VALUE = re.compile(r"['\"]([^'\"]+)['\"]")


def parse_query_statement(sql: str) -> tuple[str | None, list[str]]:
    table_match = FROM_PATTERN.search(sql or "")
    company_match = COMPANY_PATTERN.search(sql or "")
    table_name = table_match.group(1).split(".")[-1] if table_match else None
    companies = QUOTED_VALUE.findall(company_match.group(1)) if company_match else []
    return table_name, companies


class HrdService:
    def list_interfaces(
        self,
        server: PoServer,
        company_codes: list[str],
        table_names: list[str],
        search_ifid: str | None,
    ) -> list[dict]:
        inventory = ChannelService().inventory(
            server,
            component_id="*",
            channel_pattern=settings.sap_hrd_channel_pattern,
        )
        selected_companies = {value.upper() for value in company_codes}
        selected_tables = {value.upper() for value in table_names}
        needle = (search_ifid or "").strip().upper()
        rows: list[dict] = []
        for item in inventory:
            detail = ChannelService().detail(
                server,
                item["component_id"],
                item["channel_id"],
            )
            attributes = detail.get("attributes", {})
            sql = self._query_text(attributes)
            table_name, companies = parse_query_statement(sql)
            controlled = self._controlled_ifids(attributes)
            channel_id = item["channel_id"]
            if_id = self._if_id(channel_id)
            if needle and needle not in channel_id.upper() and all(
                needle not in value.upper() for value in controlled
            ):
                continue
            if selected_tables and (table_name or "").upper() not in selected_tables:
                continue
            company_set = {value.upper() for value in companies}
            if selected_companies and company_set != selected_companies:
                continue
            rows.append(
                {
                    "sid": server.sid,
                    "component_id": item["component_id"],
                    "channel_id": channel_id,
                    "if_id": if_id,
                    "dist_if_id": channel_id,
                    "company_cd": companies,
                    "dist_yn": "Y" if controlled else "N",
                    "dist_cnt": len(controlled),
                    "dist_details": [{"if_id": value, "active": True} for value in controlled],
                    "table_name": table_name,
                    "batch_tm": self._schedule_for(channel_id),
                    "data_type": str(attributes.get("dataType") or attributes.get("DataType") or ""),
                    "match_type": (
                        "EXTERNAL" if needle and any(needle in value.upper() for value in controlled)
                        else ("INTERNAL" if needle else "ALL")
                    ),
                }
            )
        return rows

    @staticmethod
    def _query_text(attributes: dict[str, Any]) -> str:
        candidates = [
            value for name, value in attributes.items()
            if "query" in name.lower() or "sql" in name.lower()
        ]
        return "\n".join(str(value) for value in candidates if value is not None)

    @staticmethod
    def _controlled_ifids(attributes: dict[str, Any]) -> list[str]:
        values: list[str] = []
        for name, value in attributes.items():
            lowered = name.lower()
            if "ifid" not in lowered and "interface" not in lowered:
                continue
            if isinstance(value, list):
                values.extend(str(item) for item in value)
            else:
                values.extend(
                    token.strip()
                    for token in re.split(r"[,;\s]+", str(value or ""))
                    if token.strip()
                )
        return list(dict.fromkeys(values))

    @staticmethod
    def _if_id(channel_id: str) -> str:
        match = re.search(r"(HRD[0-9A-Z_]+)$", channel_id, re.IGNORECASE)
        return match.group(1) if match else channel_id

    def _schedule_for(self, receiver_channel: str) -> str | None:
        if not settings.sap_hana_configured:
            return None
        sender_channel = receiver_channel.replace("JDBC4_Receiver_", "JDBC_Sender_", 1)
        try:
            from hdbcli import dbapi

            connection = dbapi.connect(
                address=settings.sap_hana_host,
                port=settings.sap_hana_port,
                user=settings.sap_hana_user,
                password=settings.sap_hana_password.get_secret_value(),
            )
            try:
                cursor = connection.cursor()
                cursor.execute(
                    """
                    select a.description
                    from xi_af_adm_schedule a, xi_af_adm_schd_scs b, xi_af_cpa_channel c
                    where substring(cast(bintostr(cast(a.channel_set as binary)) as varchar), 95, 32) = b.id
                      and cast(bintostr(cast(b.channels as binary)) as varchar) like ?
                      and c.channel = ?
                    order by a.description
                    """,
                    (f"%{sender_channel}%", sender_channel),
                )
                row = cursor.fetchone()
                return str(row[0]) if row else None
            finally:
                connection.close()
        except Exception as exc:
            raise SapPoConnectionError("SAP HANA HRD schedule lookup failed") from exc

    def send_test_message(self, server: PoServer, if_id: str) -> dict:
        sender_service = settings.sap_hrd_sender_services.get(server.sid)
        if not sender_service or not settings.sap_hrd_test_interface:
            raise SapPoConfigurationError(
                f"HRD test-message mapping is not configured for SID {server.sid}"
            )
        safe_if_id = escape(if_id)
        payload = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<ns0:MT_HRD_TEST xmlns:ns0="urn:po-monitor:hrd:test">'
            f"<DEST_IFID>{safe_if_id}</DEST_IFID>"
            "</ns0:MT_HRD_TEST>"
        )
        session = build_session(server)
        try:
            response = session.post(
                endpoint(server, settings.sap_hrd_test_path),
                params={
                    "interface": settings.sap_hrd_test_interface,
                    "senderService": sender_service,
                    "qos": "EOIO",
                },
                data=payload.encode("utf-8"),
                headers={"Content-Type": "application/xml; charset=utf-8"},
                timeout=(
                    settings.sap_connect_timeout_seconds,
                    settings.sap_read_timeout_seconds,
                ),
            )
            response.raise_for_status()
            return {
                "success": True,
                "sid": server.sid,
                "if_id": if_id,
                "http_status": response.status_code,
            }
        except Exception as exc:
            raise SapPoConnectionError("HRD test message request failed") from exc


def generate_excel(rows: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "HRD Interfaces"
    headers = [
        "I/F ID", "DIST I/F ID", "Table", "Batch", "Company",
        "DIST Count", "Match Type", "SID",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="173F5F")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(
            [
                row.get("if_id"),
                row.get("dist_if_id"),
                row.get("table_name"),
                row.get("batch_tm"),
                ", ".join(row.get("company_cd") or []),
                row.get("dist_cnt"),
                row.get("match_type"),
                row.get("sid"),
            ]
        )
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        sheet.column_dimensions[column[0].column_letter].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
