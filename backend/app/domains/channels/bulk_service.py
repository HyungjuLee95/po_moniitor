from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import PoServer
from app.domains.channels.service import ChannelService


EXPORT_FIELDS = [
    "component_id",
    "channel_id",
    "direction",
    "adapter_type",
    "dbuser",
    "dbpassword",
    "connectionURL",
]
COMPARE_FIELDS = ["dbuser", "dbpassword", "connectionURL", "adapter_type"]


class ChannelBulkService:
    def export(
        self,
        server: PoServer,
        component_id: str,
        channel_pattern: str,
    ) -> bytes:
        inventory = ChannelService().inventory(server, component_id, channel_pattern)
        rows: list[dict[str, Any]] = []
        for item in inventory:
            detail = ChannelService().detail(
                server,
                item["component_id"],
                item["channel_id"],
                include_password=True,
            )
            attributes = detail.get("attributes", {})
            rows.append(
                {
                    "component_id": item["component_id"],
                    "channel_id": item["channel_id"],
                    "direction": attributes.get("Direction") or attributes.get("direction"),
                    "adapter_type": attributes.get("Adapter Type") or attributes.get("AdapterName"),
                    "dbuser": attributes.get("dbuser"),
                    "dbpassword": detail.get("password"),
                    "connectionURL": attributes.get("connectionURL"),
                }
            )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Channels"
        sheet.append(EXPORT_FIELDS)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="173F5F")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            sheet.append([row.get(field) for field in EXPORT_FIELDS])
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 80)
            sheet.column_dimensions[column[0].column_letter].width = width
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def preview(self, server: PoServer, content: bytes) -> dict:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(values)]
        except StopIteration:
            return {"total": 0, "to_change": 0, "errors": ["빈 Excel 파일입니다."], "details": []}
        missing = {"component_id", "channel_id"} - set(headers)
        if missing:
            return {
                "total": 0,
                "to_change": 0,
                "errors": [f"필수 열 누락: {', '.join(sorted(missing))}"],
                "details": [],
            }
        details = []
        errors = []
        total = 0
        for row_number, values_row in enumerate(values, 2):
            uploaded = dict(zip(headers, values_row, strict=False))
            if not uploaded.get("component_id") or not uploaded.get("channel_id"):
                continue
            total += 1
            try:
                current_detail = ChannelService().detail(
                    server,
                    str(uploaded["component_id"]),
                    str(uploaded["channel_id"]),
                    include_password=True,
                )
                attributes = current_detail.get("attributes", {})
                current = {
                    "dbuser": attributes.get("dbuser"),
                    "dbpassword": current_detail.get("password"),
                    "connectionURL": attributes.get("connectionURL"),
                    "adapter_type": attributes.get("Adapter Type") or attributes.get("AdapterName"),
                }
                changes = [
                    field for field in COMPARE_FIELDS
                    if self._normalized(current.get(field))
                    != self._normalized(uploaded.get(field))
                ]
                if changes:
                    details.append(
                        {
                            "row": row_number,
                            "channel": f"{uploaded['component_id']}|{uploaded['channel_id']}",
                            "changes": changes,
                            "before": self._masked(current, changes),
                            "after": self._masked(uploaded, changes),
                        }
                    )
            except Exception:
                errors.append(f"{row_number}행 채널을 조회하지 못했습니다.")
        return {
            "total": total,
            "to_change": len(details),
            "errors": errors,
            "details": details,
        }

    @staticmethod
    def _normalized(value: Any) -> str:
        return str(value or "").strip()

    @staticmethod
    def _masked(values: dict, fields: list[str]) -> dict:
        return {
            field: ("********" if field == "dbpassword" and values.get(field) else values.get(field))
            for field in fields
        }
